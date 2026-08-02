import sys
from types import SimpleNamespace

import pytest

from app.services import document_parser_service as parser


def test_parse_legacy_doc_uses_antiword_with_utf8_mapping(monkeypatch, tmp_path):
    doc_path = tmp_path / "legacy.doc"
    doc_path.write_bytes(b"mock-legacy-doc")
    captured = {}

    def _mock_run(args, **kwargs):
        captured["args"] = args
        captured["kwargs"] = kwargs
        return SimpleNamespace(
            returncode=0,
            stdout="项目名称：旧版 Word Brief\n预算：30-50万".encode("utf-8"),
            stderr=b"",
        )

    monkeypatch.setattr(parser.subprocess, "run", _mock_run)

    sections = parser.parse_document(str(doc_path), "legacy.doc")

    assert captured["args"] == [
        "antiword",
        "-m",
        "UTF-8.txt",
        str(doc_path),
    ]
    assert captured["kwargs"]["check"] is False
    assert captured["kwargs"]["capture_output"] is True
    assert captured["kwargs"]["timeout"] == parser.LEGACY_DOC_PARSE_TIMEOUT_SECONDS
    assert sections == [
        parser.ParsedSection(
            label="Word正文",
            page=None,
            text="项目名称：旧版 Word Brief\n预算：30-50万",
        )
    ]


def test_parse_legacy_doc_reports_missing_antiword(monkeypatch, tmp_path):
    doc_path = tmp_path / "legacy.doc"
    doc_path.write_bytes(b"mock-legacy-doc")

    def _missing_binary(*_args, **_kwargs):
        raise FileNotFoundError

    monkeypatch.setattr(parser.subprocess, "run", _missing_binary)

    with pytest.raises(parser.DocumentParseError, match="antiword"):
        parser.parse_document(str(doc_path), "legacy.doc")


def test_parse_legacy_doc_rejects_empty_output(monkeypatch, tmp_path):
    doc_path = tmp_path / "legacy.doc"
    doc_path.write_bytes(b"mock-legacy-doc")

    monkeypatch.setattr(
        parser.subprocess,
        "run",
        lambda *_args, **_kwargs: SimpleNamespace(
            returncode=0,
            stdout=b"",
            stderr=b"",
        ),
    )

    with pytest.raises(parser.DocumentParseError, match="没有可提取的文字"):
        parser.parse_document(str(doc_path), "legacy.doc")


def test_parse_xlsx_preserves_sheet_names_and_tabular_values(tmp_path):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "项目需求"
    sheet.append(["项目名称", "夏季冰饮裸眼3D"])
    sheet.append(["投放城市", "杭州湖滨银泰"])
    workbook.create_sheet("空白页")
    excel_path = tmp_path / "brief.xlsx"
    workbook.save(excel_path)

    sections = parser.parse_document(str(excel_path), "brief.xlsx")

    assert sections == [
        parser.ParsedSection(
            label="工作表：项目需求",
            page=None,
            text="项目名称 | 夏季冰饮裸眼3D\n投放城市 | 杭州湖滨银泰",
        ),
        parser.ParsedSection(label="工作表：空白页", page=None, text=""),
    ]


def test_parse_xls_preserves_sheet_names_and_tabular_values(monkeypatch, tmp_path):
    xls_path = tmp_path / "legacy-brief.xls"
    xls_path.write_bytes(b"mock-xls")

    class _Sheet:
        nrows = 2

        @staticmethod
        def row_values(index):
            return [["预算", 300000.0], ["预计上刊时间", "2026-08-20"]][index]

    class _Workbook:
        @staticmethod
        def sheet_names():
            return ["项目需求"]

        @staticmethod
        def sheet_by_name(name):
            assert name == "项目需求"
            return _Sheet()

    fake_xlrd = SimpleNamespace(open_workbook=lambda path, on_demand: _Workbook())
    monkeypatch.setitem(sys.modules, "xlrd", fake_xlrd)

    sections = parser.parse_document(str(xls_path), "legacy-brief.xls")

    assert sections == [
        parser.ParsedSection(
            label="工作表：项目需求",
            page=None,
            text="预算 | 300000\n预计上刊时间 | 2026-08-20",
        )
    ]
